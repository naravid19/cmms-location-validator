import pandas as pd
import numpy as np
import os
import re
import logging
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

# Setup logging
# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Config:
    """Configuration constants for the application."""
    SHEET_NAME = "LTK-H"
    FILE_INPUT = r"C:\Users\narav\Desktop\Egat\other\Template MxLoader LOCATION_LTK-H.xlsm"
    DATABASE_CODE = r"C:\Users\narav\Desktop\CE code\Python\Learn\Mypython\Pandas\KKS\Database_Code.xlsx"
    
    # Column definitions
    COLS_MAIN = "A:N"
    COLS_KKS = "A:B"
    COLS_COST = "A:K"
    COLS_PLANT = "A:D"
    
    # Sheet names in Database_Code.xlsx
    SHEET_SYS = "system_code"
    SHEET_EQ = "eq_code"
    SHEET_COM = "component_code"
    SHEET_COST = "cost_center"
    SHEET_PLANT = "plant_code"

class DataLoader:
    """Handles loading and initial preprocessing of data."""
    
    @staticmethod
    def load_reference_data(db_path):
        """Loads reference data tables from the database excel."""
        logger.info(f"Step 1/7: Loading reference data from {db_path}")
        try:
            refs = {}
            refs['sys'] = pd.read_excel(db_path, sheet_name=Config.SHEET_SYS, usecols=Config.COLS_KKS)
            refs['eq'] = pd.read_excel(db_path, sheet_name=Config.SHEET_EQ, usecols=Config.COLS_KKS)
            refs['com'] = pd.read_excel(db_path, sheet_name=Config.SHEET_COM, usecols=Config.COLS_KKS)
            refs['cost'] = pd.read_excel(db_path, sheet_name=Config.SHEET_COST, usecols=Config.COLS_COST)
            refs['plant'] = pd.read_excel(db_path, sheet_name=Config.SHEET_PLANT, usecols=Config.COLS_PLANT)
            
            # Clean reference codes
            for key in ['sys', 'eq', 'com']:
                refs[key]['code'] = refs[key]['code'].str.strip().str.upper()
            
            # Clean cost center data
            cols_strip = ["Cost Center", "Name", "Description", "Hierachy Area", "Business Area", "Profit Center", "Funcional Area"]
            cols_upper = ["Cost Center", "Hierachy Area", "Business Area"]
            
            refs['cost'][cols_strip] = refs['cost'][cols_strip].apply(lambda x: x.str.strip())
            refs['cost'][cols_upper] = refs['cost'][cols_upper].apply(lambda x: x.str.upper())
            
            # Clean plant data
            refs['plant'] = refs['plant'].apply(lambda col: col.map(lambda x: x.strip().upper() if isinstance(x, str) else x))
            
            return refs
        except Exception as e:
            logger.error(f"Failed to load reference data: {e}")
            raise

    @staticmethod
    def load_input_data(file_path, sheet_name):
        """Loads the main input data."""
        logger.info(f"Step 2/7: Loading input data from {file_path}")
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=0, usecols=Config.COLS_MAIN, skiprows=[0, 2, 3, 4, 5])
            df["LOCATION"] = df["LOCATION"].str.strip()
            df["DESCRIPTION"] = df["DESCRIPTION"].str.strip()
            return df
        except Exception as e:
            logger.error(f"Failed to load input data: {e}")
            raise

class Validator:
    """Contains validation logic for Location, Codes, and Cost Centers."""

    @staticmethod
    def validate_location_format(df):
        """Checks if LOCATION matches its stripped version."""
        df_loc = df[["LOCATION"]].copy()
        df_loc["LOCATION_STRIP"] = (
            df_loc["LOCATION"]
            .str.replace("\t", "")
            .str.replace("\r", "")
            .str.replace("\n", "")
            .str.replace("\b", "")
            .str.replace(" ", "")
        )
        return np.where(df_loc["LOCATION"] != df_loc["LOCATION_STRIP"], "FALSE", "TRUE")

    @staticmethod
    def process_kks(df):
        """Processes KKS codes to extract System, EQ, and Component."""
        # Plant regex logic
        plant_list = df["LOCATION"].str.split('-', expand=True)[0].value_counts().index.tolist()
        plant_regex = "|".join([p + "-" for p in plant_list])
        plant_regex1 = "|".join([p for p in plant_list])
        
        df_clean = df.dropna(axis="index", how="all").copy()
        
        # Remove plant prefix
        df_clean["LOCATION_x"] = df_clean["LOCATION"].str.replace(plant_regex, "", regex=True)
        df_clean["LOCATION_x"] = df_clean["LOCATION_x"].str.replace(plant_regex1, "", regex=True)
        
        # Filter valid KKS
        df_kks = df_clean[["LOCATION", "LOCATION_x", "DESCRIPTION"]].copy()
        df_kks["DESCRIPTION"] = df_kks["DESCRIPTION"].str.strip()
        df_kks["LOCATION"] = df_kks["LOCATION"].str.strip()
        df_kks["LOCATION_x"] = df_kks["LOCATION_x"].str.strip()
        
        df_kks_test = df_kks.dropna().copy()
        
        # Remove prefix pattern (e.g. 10, 11)
        lst = df_kks_test["LOCATION_x"].str[0:3].value_counts().index
        filtered_lst = [x for x in lst if re.match(r"^[A-Za-z][A-Za-z0-9]{0,2}$", x)]
        regex_pattern = "|".join(filtered_lst)
        
        replace_first = lambda x: re.sub(f"({regex_pattern})", "", x, count=1)
        df_kks_test["LOCATION_y"] = df_kks_test["LOCATION_x"].apply(replace_first)
        
        # Extract System, EQ
        system_eq = df_kks_test["LOCATION_y"].str.findall("[A-Z,-]+").str.join("")
        df_kks_test["system_eq"] = system_eq
        df_kks_test["SYSTEM"] = df_kks_test["system_eq"].str[0:3].str.upper().str.extract("([A-Z]+)", expand=False)
        df_kks_test["EQ"] = df_kks_test["system_eq"].str[3:5].str.extract("([A-Z]+)", expand=False).str.upper()
        
        def extract_component(system_eq):
            if "-" in system_eq and len(system_eq) == 7:
                return system_eq[5:].upper()
            else:
                return system_eq[5:].upper()

        df_kks_test["COMPONENT"] = df_kks_test["system_eq"].apply(extract_component)
        
        # Handle duplicates logic for DESCRIPTION_new
        # Note: We return the dataframe BEFORE dropping duplicates if we want to track them, 
        # but the original code drops them. We will follow the original flow.
        
        # Identify duplicates (for main df logic later if needed, but here we process unique KKS)
        duplicated_indices = df_kks_test[df_kks_test.duplicated()].index
        df_kks_test = df_kks_test.drop_duplicates()
        
        # Logic for DESCRIPTION_new
        index_more = (
            df_kks_test["DESCRIPTION"]
            .value_counts()[df_kks_test["DESCRIPTION"].value_counts() > 1]
            .index
        )
        df_kks_test["DESCRIPTION_new"] = ""
        for des in index_more:
            df_sub = df_kks_test[df_kks_test["DESCRIPTION"] == des].copy()
            df_sub["DESCRIPTION_new"] = df_sub["DESCRIPTION"] + "_" + df_sub["LOCATION_x"]
            df_kks_test.loc[df_sub.index, "DESCRIPTION_new"] = df_sub["DESCRIPTION_new"]
            
        return df_kks_test, duplicated_indices

    @staticmethod
    def validate_codes(df_main, df_kks_test, refs):
        """Validates System, EQ, and Component codes."""
        # Map KKS data back to main df
        df_main["SYSTEM"] = df_kks_test["SYSTEM"]
        df_main["EQ"] = df_kks_test["EQ"]
        df_main["COMPONENT"] = df_kks_test["COMPONENT"]
        
        def check_code(row, col, ref_df):
            val = row.get(col)
            if pd.isna(val) or val == "":
                return ""
            return "มี" if val in ref_df["code"].values else "ไม่มี"

        df_main["SYSTEM_STATUS"] = df_kks_test.apply(lambda r: check_code(r, "SYSTEM", refs['sys']), axis=1)
        df_main["EQ_STATUS"] = df_kks_test.apply(lambda r: check_code(r, "EQ", refs['eq']), axis=1)
        df_main["COMPONENT_STATUS"] = df_kks_test.apply(lambda r: check_code(r, "COMPONENT", refs['com']), axis=1)
        return df_main

    @staticmethod
    def validate_cost_center(df_original, df_cost_ref):
        """Validates cost center logic."""
        # Prepare working dataframe
        df1 = df_original.dropna(axis="index", how="all")
        df_cost = df1[["LOCATION", "EGCOSTCENTER", "EGBA", "LOCHIERARCHY.PARENT"]].copy()
        
        # Determine Plant Unit
        plant_list = df_original["LOCATION"].str.split('-', expand=True)[0].value_counts().index.tolist()
        plant_unit = len(plant_list[0]) if plant_list else 3
        
        plant_regex = "|".join([p + "-" for p in plant_list])
        plant_regex1 = "|".join([p for p in plant_list])
        df1_loc_x = df1["LOCATION"].str.replace(plant_regex, "", regex=True).str.replace(plant_regex1, "", regex=True)
        
        df_cost["TOTAL_PLANT"] = df1_loc_x.str[:3]
        df_cost["NUM_PLANT"] = ''
        
        cond1_num_plant = df_original['LOCATION'].isna()
        cond2_num_plant = (df_cost["TOTAL_PLANT"] == '') | (df_cost["TOTAL_PLANT"].isna())
        
        df_cost.loc[cond1_num_plant, "NUM_PLANT"] = 'ไม่มี LOCATION'
        df_cost.loc[cond2_num_plant, "NUM_PLANT"] = 'Common'
        df_cost.loc[~cond2_num_plant, "NUM_PLANT"] = df_cost["TOTAL_PLANT"]
        
        df_cost.loc[df_cost["LOCATION"].isna(), "COST_STATUS"] = 'ไม่มี LOCATION'
        
        # Preprocess ref
        def preprocess_plant_names(df_make_cost):
            df_make_cost["Plant Name Split"] = df_make_cost["Plant Name"].apply(lambda x: [name.strip() for name in x.split(',')] if ',' in x else [x.strip()])
            df_make_cost["Plant Name1 Split"] = df_make_cost["Plant Name1"].apply(lambda x: [name.strip() for name in x.split(',')] if ',' in x else [x.strip()])
            return df_make_cost

        df_cost_ref["Plant Name"] = df_cost_ref["Plant Name"].astype(str)
        df_cost_ref["Plant Name1"] = df_cost_ref["Plant Name1"].astype(str)
        df_cost_ref = preprocess_plant_names(df_cost_ref)
        
        df_cost["NUM_PLANT1"] = df_cost["NUM_PLANT"]
        
        # Update extracted numbers logic
        def update_extracted_numbers(row, df_make_cost, plant_unit):
            if plant_unit not in [2, 3, 4]:
                raise ValueError("Invalid value for plant_unit. Only 2, 3 or 4 are allowed.")
            
            prefix_length = plant_unit
            total_plant_prefix = row["LOCATION"][:prefix_length] if isinstance(row["LOCATION"], str) else ""
            
            cost_center_row = df_make_cost[
                df_make_cost["Plant Name Split"].apply(lambda x: total_plant_prefix in x) |
                df_make_cost["Plant Name1 Split"].apply(lambda x: total_plant_prefix in x)
            ]
            if not cost_center_row.empty:  
                plant_unit_values = cost_center_row['Plant Unit'].dropna().astype(str).str.split(',').explode().str.strip()
                plant_unit_values1 = cost_center_row['Plant Unit1'].dropna().astype(str).str.split(',').explode().str.strip()
                
                if cost_center_row["Plant Name Split"].apply(lambda x: total_plant_prefix in x).any():
                    if row["NUM_PLANT1"] not in plant_unit_values.values:
                        row["NUM_PLANT1"] = 'ไม่พบ Plant Unit'
                elif cost_center_row["Plant Name1 Split"].apply(lambda x: total_plant_prefix in x).any():
                    if row["NUM_PLANT1"] not in plant_unit_values1.values:
                        row["NUM_PLANT1"] = 'ไม่พบ Plant Unit'
            elif row["NUM_PLANT1"] == 'ไม่มี LOCATION':
                return row
            else:
                row["NUM_PLANT1"] = 'ไม่พบ Plant Name'
            return row

        df_cost = df_cost.apply(lambda row: update_extracted_numbers(row, df_cost_ref, plant_unit), axis=1)
        
        if 'COST_SHOULD_BE' not in df_cost.columns:
            df_cost['COST_SHOULD_BE'] = ''

        def cost_center_check(row, df_make_cost, plant_unit):
            if plant_unit not in [2, 3, 4]:
                raise ValueError("Invalid value for plant_unit. Only 2, 3 or 4 are allowed.")
            prefix_length = plant_unit
            total_plant_prefix = row["LOCATION"][:prefix_length] if isinstance(row["LOCATION"], str) else ""
            if pd.isna(row["LOCATION"]):
                return 'ไม่มี LOCATION'
            
            matching_row = df_make_cost[
                df_make_cost["Plant Name Split"].apply(lambda x: total_plant_prefix in x) |
                df_make_cost["Plant Name1 Split"].apply(lambda x: total_plant_prefix in x)
            ]
            
            if matching_row.empty and not pd.isna(row["LOCATION"]):
                df_cost.at[row.name, 'COST_SHOULD_BE'] = 're_check'
                return row["NUM_PLANT1"]
            
            plant_unit_values = matching_row['Plant Unit'].dropna().astype(str).str.split(',').explode().str.strip()
            plant_unit_values1 = matching_row['Plant Unit1'].dropna().astype(str).str.split(',').explode().str.strip()
            
            if (total_plant_prefix in matching_row["Plant Name Split"].explode().values and
                row['NUM_PLANT1'] in plant_unit_values.values):
                matched_row = matching_row[matching_row.apply(lambda x: row['NUM_PLANT1'] in str(x['Plant Unit']).split(','), axis=1)]
            elif (total_plant_prefix in matching_row["Plant Name1 Split"].explode().values and
                row['NUM_PLANT1'] in plant_unit_values1.values):
                matched_row = matching_row[matching_row.apply(lambda x: row['NUM_PLANT1'] in str(x['Plant Unit1']).split(','), axis=1)]
            else:
                matched_row = pd.DataFrame()
            
            if matched_row.empty and not pd.isna(row["LOCATION"]):
                df_cost.at[row.name, 'COST_SHOULD_BE'] = 're_check'
                return 'ไม่พบ Plant Unit'
            elif matched_row.empty:
                return 'ข้อผิดพลาดใหม่'
            
            cost_center_match = row['EGCOSTCENTER'] == matched_row.iloc[0]['Cost Center']
            business_area_match = row['EGBA'] == matched_row.iloc[0]['Business Area']
            
            if not cost_center_match and not business_area_match:
                if pd.isna(row['EGCOSTCENTER']) and pd.isna(row['EGBA']):
                    df_cost.at[row.name, 'COST_SHOULD_BE'] = f"{matched_row.iloc[0]['Cost Center']},{matched_row.iloc[0]['Business Area']}"
                    return 'ไม่มี EGCOSTCENTER เเละ EGBA'
                df_cost.at[row.name, 'COST_SHOULD_BE'] = f"{matched_row.iloc[0]['Cost Center']},{matched_row.iloc[0]['Business Area']}"
                return 'EGCOSTCENTER เเละ EGBA ไม่สอดคล้องกัน'
            elif not cost_center_match:
                if pd.isna(row['EGCOSTCENTER']):
                    df_cost.at[row.name, 'COST_SHOULD_BE'] = matched_row.iloc[0]['Cost Center']
                    return 'ไม่มี EGCOSTCENTER'
                
                if  ('Common' in plant_unit_values.values or 
                    'Common' in plant_unit_values1.values):
                    cost_center = matched_row.iloc[0]['Cost Center']
                    modified_cost_center = (cost_center + '00' if len(cost_center) == 7 
                                            else cost_center[:-2] if len(cost_center) == 9 and cost_center.endswith('00')
                                            else cost_center)
                    
                    if row['EGCOSTCENTER'] == modified_cost_center:
                        df_cost.at[row.name, 'COST_SHOULD_BE'] = 'do_nothing'
                        return 'OK'
                    else:
                        df_cost.at[row.name, 'COST_SHOULD_BE'] = modified_cost_center
                        return 'EGCOSTCENTER ไม่สอดคล้องกัน'
                else:
                    df_cost.at[row.name, 'COST_SHOULD_BE'] = matched_row.iloc[0]['Cost Center']
                    return 'EGCOSTCENTER ไม่สอดคล้องกัน'
            elif not business_area_match:
                if pd.isna(row['EGBA']):
                    df_cost.at[row.name, 'COST_SHOULD_BE'] = matched_row.iloc[0]['Business Area']
                    return 'ไม่มี EGBA'
                df_cost.at[row.name, 'COST_SHOULD_BE'] = matched_row.iloc[0]['Business Area']
                return 'EGBA ไม่สอดคล้องกัน'
            
            df_cost.at[row.name, 'COST_SHOULD_BE'] = 'do_nothing'
            return 'OK'

        df_cost['COST_STATUS'] = df_cost.apply(lambda row: cost_center_check(row, df_cost_ref, plant_unit), axis=1)
        return df_cost

    @staticmethod
    def validate_parent(df_original):
        """Validates parent hierarchy."""
        df_parent = df_original[["LOCATION", "LOCHIERARCHY.PARENT"]].copy()
        df_parent["PARENT_STATUS"] = ''
        
        non_na_condition = df_parent["LOCATION"].notna() & df_parent["LOCHIERARCHY.PARENT"].notna()

        df_parent.loc[non_na_condition, "PARENT_STATUS"] = df_parent.loc[non_na_condition].apply(
            lambda row: ('OK' if df_parent["LOCATION"].eq(row["LOCHIERARCHY.PARENT"]).any() else 'ไม่พบ PARENT') 
            if pd.notna(row["LOCATION"]) and pd.notna(row["LOCHIERARCHY.PARENT"]) and row["LOCHIERARCHY.PARENT"] in row["LOCATION"] 
            else 'PARENT ไม่สอดคล้อง', axis=1
        )

        df_parent.loc[df_parent["LOCATION"].isna(), "PARENT_STATUS"] = 'ไม่มี LOCATION'
        df_parent.loc[df_parent["LOCHIERARCHY.PARENT"].isna(), "PARENT_STATUS"] = 'ไม่มี PARENT'
        
        return df_parent["PARENT_STATUS"]

class ExcelReporter:
    """Handles formatting and saving the output Excel."""
    
    @staticmethod
    def generate_excel_report(file_input, sheet_name, file_output, ws1_data_source_file):
        """
        Applies the exact formatting logic from the original script.
        """
        logger.info(f"Step 7/7: Generating Excel report: {file_output}")
        
        try:
            wb1 = openpyxl.load_workbook(file_input, keep_vba=False, data_only=False)
            ws1 = wb1[sheet_name]
        except Exception as e:
            logger.error(f"Error loading workbook 1: {e}")
            exit(1)

        try:
            wb2 = load_workbook(ws1_data_source_file)
            ws2 = wb2.active
        except Exception as e:
            logger.error(f"Error loading workbook 2: {e}")
            exit(1)

        # Delete all sheets except the one specified
        all_sheets = wb1.sheetnames
        for sheet in all_sheets:
            if sheet != sheet_name:
                del wb1[sheet]

        blue_fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        ws1.insert_cols(1,8)    # insert_cols 8 columns

        row_offset = 6  # Start writing at row 7 in ws1
        col_offset = 1  # Start writing at column A in ws1

        for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=2, max_col=9)):   # B1 : I1
            for j, cell in enumerate(row):
                target_row = i + row_offset
                target_col = j + col_offset
                ws1.cell(row=target_row, column=target_col, value=cell.value)

        # Check Location and Description
        for row in range(7, ws1.max_row + 1):   # Start from row 7
            cell = ws1.cell(row=row, column=5)  # Column E
            fill = None
            if cell.value == 1:
                fill = yellow_fill
            elif cell.value == 2:
                fill = red_fill
            
            if fill:
                for col in range(1, 6): # A to E
                    ws1.cell(row=row, column=col).fill = fill

        # Check COST_STATUS
        for row in range(7, ws1.max_row + 1):
            cell = ws1.cell(row=row, column=6)  # Column F
            fill = None
            if cell.value == '':
                fill = yellow_fill
            elif cell.value in ['ไม่มี LOCATION', 'EGCOSTCENTER เเละ EGBA ไม่สอดคล้องกัน', 'EGCOSTCENTER ไม่สอดคล้องกัน', 
                                'EGBA ไม่สอดคล้องกัน', 'ไม่มี EGCOSTCENTER เเละ EGBA', 'ไม่มี EGCOSTCENTER', 
                                'ไม่มี EGBA', 'ไม่พบ Plant Name', 'ไม่พบ Plant Unit', 'ข้อผิดพลาดใหม่']:
                fill = red_fill
            elif cell.value == 'OK':
                fill = None
            else:
                fill = blue_fill
            
            if fill:
                cell.fill = fill
        
        # Check PARENT_STATUS
        for row in range(7, ws1.max_row + 1):
            cell = ws1.cell(row=row, column=8) # Column H
            fill = None
            if cell.value == 'ไม่พบ PARENT':
                fill = yellow_fill
            elif cell.value in ['PARENT ไม่สอดคล้อง', 'ไม่มี LOCATION', 'ไม่มี PARENT']:
                fill = red_fill
            elif cell.value == 'OK':
                fill = None
            else:
                fill = blue_fill
            
            if fill:
                cell.fill = fill

        # Check the accuracy of the LOCATION
        for row in range(2, ws2.max_row + 1):
            value_in_I = ws2.cell(row=row, column=10).value  # Column J is the 9th column
            value_in_B = ws2.cell(row=row, column=2).value  # Column B is the 2nd column

            if value_in_I == 'FALSE':
                # Check if value in ws2 column B row matches any value in ws1 column H starting from row 7
                for row_ws1 in range(7, ws1.max_row + 1):
                    if ws1.cell(row=row_ws1, column=8).value == value_in_B:  # Column H is the 8th column
                        ws1.cell(row=row_ws1, column=8).fill = yellow_fill

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            for row in range(1, 7):
                cell = ws1[f'{col}{row}']
                cell.fill = blue_fill

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, 9):
            cell = ws1.cell(row=6, column=col)
            cell.border = thin_border

        ws1.auto_filter.ref = "A6:V6"

        column_widths = {
            'F': 29.78, 'G': 29.78, 'H': 29.78, 'I': 23.33, 'J': 30.56, 'K': 15.56,
            'L': 22.0, 'M': 19.89, 'N': 19.89, 'O': 19.89, 'P': 19.89, 'Q': 26.89,
            'R': 26.89, 'S': 18.67, 'T': 19.89, 'U': 26.22, 'V': 15.67,
        }
        for col, width in column_widths.items():
            ws1.column_dimensions[col].width = width

        try:
            wb1.save(file_output)
            logger.info(f"Successfully saved report to {file_output}")
        except Exception as e:
            logger.error(f"Error saving file: {e}")
        finally:
            wb1.close()
            wb2.close()

def main(sheet_name=None, file_input=None, database_code=None):
    logger.info("=== Starting Location Validator v1.0.0 ===")
    
    # Update Config if arguments are provided
    if sheet_name:
        Config.SHEET_NAME = sheet_name
    if file_input:
        Config.FILE_INPUT = file_input
    if database_code:
        Config.DATABASE_CODE = database_code

    logger.info(f"Configuration: Sheet={Config.SHEET_NAME}, Input={Config.FILE_INPUT}, DB={Config.DATABASE_CODE}")

    # 1. Load Data
    try:
        refs = DataLoader.load_reference_data(Config.DATABASE_CODE)
        df_main = DataLoader.load_input_data(Config.FILE_INPUT, Config.SHEET_NAME)
    except Exception as e:
        logger.critical(f"Initialization failed: {e}")
        return False # Return failure

    # 2. Validate Location Format
    logger.info("Step 3/7: Validating Location Format...")
    try:
        df_main["LOCATION_STATUS"] = Validator.validate_location_format(df_main)
        
        # 3. Process KKS for Codes
        logger.info("Step 4/7: Processing KKS Codes...")
        df_kks_test, duplicated_indices = Validator.process_kks(df_main)
        
        # 4. Validate Codes (System, EQ, Component)
        # Map DESCRIPTION_new back to df_main
        df_main["DESCRIPTION_new"] = ""
        df_main.loc[df_kks_test.index, "DESCRIPTION_new"] = df_kks_test["DESCRIPTION_new"]
        
        # Apply logic for COMMENT, SHOULD_BE, LEVEL
        all_na_rows = df_main.isna().all(axis=1)
        df_main.loc[all_na_rows] = df_main.loc[all_na_rows].astype(object)
        df_main.loc[all_na_rows] = df_main.loc[all_na_rows].fillna("xx")
        df_main["COMMENT"] = ""
        df_main["SHOULD_BE"] = ""
        df_main["LEVEL"] = 0
        
        # Handle duplicates logic
        if not duplicated_indices.empty:
            df_main.loc[duplicated_indices, "COMMENT"] = "kks และ description ซ้ำกับแถวอื่นๆ"
            df_main.loc[duplicated_indices, "SHOULD_BE"] = "ลบทิ้ง"
            df_main.loc[duplicated_indices, "LEVEL"] = 2
        
        index_location = df_main[df_main["LOCATION"].isna()].index
        df_main.loc[index_location, "COMMENT"] = "ไม่พบ kks location"
        df_main.loc[index_location, "SHOULD_BE"] = "re_check"
        df_main.loc[index_location, "LEVEL"] = 2
        
        index_desc = df_main[df_main["DESCRIPTION"].isna()].index
        df_main.loc[index_desc, "COMMENT"] = "ไม่พบ description"
        df_main.loc[index_desc, "SHOULD_BE"] = "re_check"
        df_main.loc[index_desc, "LEVEL"] = 2
        
        cond1 = (df_main["DESCRIPTION_new"] == "")
        cond2 = df_main["LOCATION"].notna()
        cond3 = df_main["DESCRIPTION"].notna()
        index_null_desc = df_main[cond1 & cond2 & cond3].index
        df_main.loc[index_null_desc, "COMMENT"] = "Ok"
        df_main.loc[index_null_desc, "SHOULD_BE"] = "do_nothing"
        df_main.loc[index_null_desc, "LEVEL"] = 0
        
        cond1 = ((df_main["DESCRIPTION_new"] != "") & (df_main["DESCRIPTION_new"] != "xx") & (df_main["DESCRIPTION_new"].notna()))
        index_not_null_desc = df_main[cond1 & cond2 & cond3].index
        
        df_main.loc[index_not_null_desc, "COMMENT"] = "description ซ้ำกันแต่ kks ไม่ซ้ำ"
        df_main.loc[index_not_null_desc, "SHOULD_BE"] = df_main.loc[index_not_null_desc, "DESCRIPTION_new"]
        df_main.loc[index_not_null_desc, "LEVEL"] = 1

        # Validate Codes
        logger.info("Step 5/7: Validating System, EQ, and Component Codes...")
        df_main = Validator.validate_codes(df_main, df_kks_test, refs)

        # 5. Validate Cost Center
        logger.info("Step 6/7: Validating Cost Centers and Hierarchy...")
        df_cost = Validator.validate_cost_center(df_main, refs['cost'])
        df_main["COST_STATUS"] = df_cost["COST_STATUS"]
        df_main["COST_SHOULD_BE"] = df_cost["COST_SHOULD_BE"]

        # 6. Validate Parent
        df_main["PARENT_STATUS"] = Validator.validate_parent(df_main)

        # 7. Generate Output
        output_cols = [
            "LOCATION", "DESCRIPTION", "COMMENT", "SHOULD_BE", "LEVEL",
            "COST_STATUS", "COST_SHOULD_BE", "PARENT_STATUS", "LOCATION_STATUS",
            "SYSTEM", "SYSTEM_STATUS", "EQ", "EQ_STATUS", "COMPONENT", "COMPONENT_STATUS"
        ]
        
        for col in output_cols:
            if col not in df_main.columns:
                df_main[col] = ""
                
        # Save intermediate file
        file_dir = os.path.dirname(Config.FILE_INPUT)
        file_output_name = f"Location_review_{Config.SHEET_NAME}.xlsx"
        file_intermediate = os.path.join(file_dir, file_output_name)
        
        df_main[output_cols].to_excel(file_intermediate)
        
        # Generate Final Report
        file_base = os.path.splitext(os.path.basename(Config.FILE_INPUT))[0]
        file_final = os.path.join(file_dir, f"{file_base}(REVIEW).xlsx")
        
        ExcelReporter.generate_excel_report(Config.FILE_INPUT, Config.SHEET_NAME, file_final, file_intermediate)
        logger.info("=== Processing Complete Successfully ===")
        return True # Return success

    except Exception as e:
        logger.error(f"An error occurred during processing: {e}")
        return False

if __name__ == "__main__":
    main()
